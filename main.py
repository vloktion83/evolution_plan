# задача кода:
#     получить теущие заявки
#     получить текущие остатки
#     получить текущие планы производства
#     учесть огрничения в ресурсах и подобрать оптимальный план призводства

# ОГРАНИЧЕНИЯ:
# код не учитывает есть ли на складе назафасованная продукция. Возможна ситуация, когда нужно не произвести,
# а зафасовать, Но это повлечет за собой в логике программы ограничения производственных ресурсов.
import os
import subprocess
# загрузка необходимых библиотек
import inspect
import pickle
import logging
import gzip
import traceback
from os import getpid
from socket import gethostname
import argparse
import numpy as np
# import cupy as np
import pandas as pd
import os as os
import math
import traceback
import time
import datetime
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import random
random.seed()
import copy
import requests
# отключение уведомлений о небольших ошибках
import warnings
warnings.filterwarnings('ignore')
import multiprocessing

from prettytable import PrettyTable

pd.set_option('display.max_rows', 550)
pd.set_option('display.max_columns', 150)

# Устанавливаем формат отображения для чисел в numpy
np.set_printoptions(precision=5, suppress=True)

# Устанавливаем формат отображения для чисел в pandas
pd.options.display.float_format = '{:.2f}'.format
pd.options.display.float_format = '{:,.0f}'.format

# настройка бота в телеграм
TOKEN = "6439740950:AAHawPVBYPHs5nwZEIkBv3weBjk2UxqgTX4"
chat_id = "-1001917632339"

def send_message(message):
    # отправляет отладочную информацию в телеграм канал
    # message = "Здесь напишите свое сообщение"
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage?chat_id={chat_id}&text={message}"
    requests.get(url).json() # Эта строка отсылает сообщение

def send_file(file):
    files = {'document': open(file, 'rb')}
    requests.post(f'https://api.telegram.org/bot{TOKEN}/sendDocument?chat_id={chat_id}', files=files)

import requests

def send_media_group(files):
    media = []
    for file_name in files:
        response = requests.post(
            f'https://api.telegram.org/bot{TOKEN}/sendPhoto?chat_id={chat_id}',
            files={'photo': open(file_name, 'rb')}
        )


def main():
    # парсер аргументов при запуске файла из командной строки
    parser = argparse.ArgumentParser(description="Пример использования аргументов командной строки.")

    # Аргумент для указания, что использовать только один процесс
    parser.add_argument("--no_cpucount", action="store_true", help="Использовать только один процесс")

    # Аргумент для указания времени работы каждого процесса
    parser.add_argument("--hours_to_work", type=int, choices=range(1, 65),
                         help="Время работы каждого процесса (от 1 до 65 часов)")
    # Аргумент для разрешения отправлять отладочную информацию в telegram
    parser.add_argument("--telegram", action="store_true", help='при наличии флага будет скидывать сообщения в телеграмм канал')

    # аргумент Время завершения
    parser.add_argument("--end_time", type=int, help="Точное время завершения в формате часов")

    # Аргумент для указания, что это главный процесс
    parser.add_argument("--master", action="store_true", help="Это главный процесс. без флага работает на час меньше и")

    # Аргумент для указания рабочей директории
    parser.add_argument("--work_dir", type=str, help="Рабочая директория")

    # Аргумент для указания имени файла для планирования
    parser.add_argument("--plan_name", type=str, help="Имя файла для планирования")

    # Добавляем аргумент no_change_days с значением по умолчанию 1
    parser.add_argument("--no_change_days", type=int, choices=range(0, 11), default=1,
                        help="Количество дней (от 1 до 10) без изменений")

    parser.add_argument("--no_exchange", type=int, choices=range(0, 300), default=0,
                        help="Сколько минут не будут загружаться параллельные генерации")

    args = parser.parse_args()

    return args

def master_slave(master):
    if master:
        return 'Master'
    else:
        return 'Slave'

def get_cpunum(no_cpucount):
    if no_cpucount:
        return 1
    else:
        return os.cpu_count()-1

def load_excel(file_path):
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names

    dataframes = {}

    for sheet_name in sheet_names:
        dataframes[sheet_name] = xls.parse(sheet_name).fillna(0)

    return dataframes

def dict_to_globals(dataframes):
    # Создание функции для конвертации словаря в глобальные переменные
    for key in dataframes.keys():
        globals()[key] = dataframes[key]

# функция принимает на вход первую и последню дату встреченную в входных данных
def create_date_columns(start_date, last_date):
    desired_format = '%d-%m-%Y'
    date_columns = []
    new_day = start_date
    while new_day != last_date + timedelta(days=1):
        date_columns.append(new_day)
        new_day = new_day + timedelta(days=1)

    formatted_dates = [date.strftime(desired_format) for date in date_columns]

    return formatted_dates

# заполняет таблицу "потребность в ресурсах" из перечисленных ресурсов и дат
def initialize_resource_need(resources, start_date, last_date):
    df = pd.DataFrame()
    df["Рабочий центр"] = resources

    date_cols = create_date_columns(start_date, last_date)
    for col in date_cols:
        df[col] = None

    return df

# Функция форматирования даты
def format_date(date_str):

    date_str = date_str

    try:
        date_obj = pd.to_datetime(date_str, format='%d.%m.%Y %H:%M:%S')

    except ValueError:

        try:
            date_obj = pd.to_datetime(date_str, format='%Y-%m-%d %H:%M:%S')

        except ValueError:
            return date_str

    if date_obj is not None:
        return date_obj.strftime('%d-%m-%Y')
    else:
        return date_str

args=main()

# Функция добавления недостающих дат
def add_missing_dates(df, dates):
    dates=set(dates)
    start_cols = df.columns.difference(dates)
    existing = set(set(df.columns)-set(start_cols))
    missing = dates - existing
    updated = list(existing)+list(missing)
    df = df.reindex(columns=list(start_cols)+list(formatted_dates)).fillna(0)
    return df


# Заполняем словарь значениями из DataFrame production_plan и ордерс
def fill_product_dict(df, nomenclature):
    product_dict = {}

    for name in nomenclature:
        try:
            product_type = df[df['Номенклатура'] == name]['Вид производимой продукции'].iloc[0]
            product_dict[name] = product_type
        except:
            pass

    return product_dict

# Функция для добавления недостающих строк с номенклатурой
def add_missing_rows(df, nomenclature_list):
    existing_nomenclature = set(df['Номенклатура'].values)
    missing_nomenclature = list(set(nomenclature_list) - existing_nomenclature)
    new_rows = [{'Номенклатура': nomenclature,'Вид производимой продукции':nomenclature_to_production.get(nomenclature, 0)}
                for nomenclature in missing_nomenclature]
    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    return df

# Заполняем остатки в будущих днях в сооствествии с текущими остатками, планами производства и заказами
# т.е. последний день в таблице не будем плинировать производство
# Цикл по датам, не берем в расчет последнюю дату, так как невозмодно посчитать остатки для даты на день позже

# процедура которая получает датафреймы с остатками, заказами и планом производства
# и отталкиваясь от начального остатка рассчитывает остатки для всех дат с уветом заказов и производства


def count_stock(stock, orders, production_plan, dates):
    for date_i in range(len(dates)-1):
        # Получаем текущие заказы, остатки и планы производства для данной даты
        current_orders = np.array(orders[dates[date_i]].tolist())

        current_stock = np.array(stock[dates[date_i]].tolist())
        # Проверяем, есть ли прошлый план производства для данной даты
        days_per_freez=1  # сколько дней на заморозку
        if date_i >= days_per_freez:
            current_plan = np.array(
                production_plan[dates[date_i-days_per_freez]].tolist())

        # Если прошлого плана нет, заполняем его нулями
        else:
            current_plan = np.array([0] * len(current_orders))
        # если есть отрицательный остаток, то для расчета остатков завтрашнего дня примем его равным 0
        current_stock = np.where(current_stock < 0, 0, current_stock)
        # Вычисляем новые остатки на следующую дату
        stock_i = current_stock-current_orders+current_plan

        # Проверяем, не является ли текущая дата последней
        if dates[date_i] != dates.iloc[-1]:

            # Обновляем остатки на следующую дату
            stock[dates[date_i+1]] = stock_i
    return stock

def ABC_XYZ(week_history: object) -> object:
    #     определяем коэффициент запасов в зависимости от категории
    abcxyz_to_coef = {'AX': 1.5, 'BX': 1.3, 'CX': 1 ,
                      'AY': 1.3 ,'BY': 1.2 ,'CY': 0 ,
                      'AZ': 0.2 ,'BZ': 0, 'CZ': 0}

    col_num = week_history.iloc[:, 2:].count().count()
    # Создание новой колонки с суммой продаж
    week_history['Total_Sales'] = week_history.iloc[:, 2:].sum(axis=1)

    # Сортировка датафрейма по сумме продаж
    week_history_sorted = week_history.sort_values(
        by='Total_Sales', ascending=False)

    # Добавление колонки с нарастающим итогом
    week_history_sorted['Cumulative_Sales'] = week_history_sorted['Total_Sales'].cumsum()

    # Создание словаря для хранения XYZ групп
    xyz_groups = {}
    name_to_mean = {}

    # Проход по каждой строке (номенклатуре) в датафрейме week_history
    for index, row in week_history_sorted.iterrows():
        nomenclature = row["Номенклатура"]
        weekly_sales = row[2:col_num]  # Продажи за каждую из 10 недель

        # Находим коэффициент вариации
        mean_sales = np.mean(weekly_sales)
        std_deviation = np.std(weekly_sales)
        coefficient_of_variation = (std_deviation / mean_sales) * 100
        # Определяем XYZ группу
        if coefficient_of_variation < 50:
            xyz_group = "X"
        elif coefficient_of_variation < 80:
            xyz_group = "Y"
        else:
            xyz_group = "Z"

        xyz_groups[nomenclature] = xyz_group
        name_to_mean[nomenclature] = mean_sales

    # Определение ABC группы
    total_sales_sum = week_history_sorted['Total_Sales'].sum()
    total_sales_threshold_A = total_sales_sum * 0.8
    total_sales_threshold_B = total_sales_sum * 0.95

    abc_groups = {}

    for index, row in week_history_sorted.iterrows():
        cumulative_sales = row['Cumulative_Sales']
        nomenclature = row['Номенклатура']

        if cumulative_sales < total_sales_threshold_A:
            abc_group = 'A'
        elif cumulative_sales < total_sales_threshold_B:
            abc_group = 'B'
        else:
            abc_group = 'C'

        abc_groups[nomenclature] = abc_group
    abcxyz = {}
    for name in abc_groups:
        abcxyz[name] = abc_groups[name] + xyz_groups[name]

    week_history_sorted['XYZ Группа'] = week_history_sorted['Номенклатура'].map(xyz_groups)
    week_history_sorted['ABC Группа'] = week_history_sorted['Номенклатура'].map(abc_groups)

    week_history_sorted['mean_sell'] = week_history_sorted['Номенклатура'].map(name_to_mean)
    week_history_sorted['ABC_XYZ'] = week_history_sorted['Номенклатура'].map(abcxyz)
    week_history_sorted['stock-coef'] = week_history_sorted['ABC_XYZ'].map(abcxyz_to_coef)
    week_history_sorted['stock-coef']=week_history_sorted['stock-coef']
    week_history_sorted['transit_stock'] = (1+week_history_sorted['stock-coef']) * week_history_sorted['mean_sell']
    week_history_sorted['Вид производимой продукции']=week_history_sorted['Номенклатура'].map(nomenclature_to_production)

    return week_history_sorted

# функция, которая будет возвращать тип дня (рабочий или выходной) на основе переданной даты

def get_day_type(date_str):
    date_obj = datetime.strptime(date_str, '%d-%m-%Y')
    day_of_week = date_obj.weekday()  # Возвращает день недели (0 - понедельник, 6 - воскресенье)

    return day_of_week


def calculate_fitness_equipment_overloads(wc_resource, resource_need):
    # возвращает штраф за превышение испошльзования мощности рабочих центров
    # штраф = превышение в процентах*превышение в килограммах
    # при превышении в процентах на 20% - увеличивается в 5 раз

    total_penalty = 0
    # if get_day_type(dates[0])==4:
    #     no_estimate_days=4
    # elif get_day_type(dates[0])==5:
    #     no_estimate_days=3
    # else:
    no_estimate_days=args.no_change_days
    capacity = np.array(wc_resource.iloc[:, 1+no_estimate_days:])
    capacity = np.where(capacity == 0, 0.0001, capacity)
    resource_need_values = np.array(resource_need.iloc[:, 1+no_estimate_days:])

    overload = (resource_need_values - capacity)
    overload = np.where(overload < 0, 0, overload)
    overload_percent = ((overload)) / capacity * 100

    # overload_percent = np.where(overload_percent >= 30, overload_percent * 50, overload_percent)
    # overload_percent = np.where((30 > overload_percent)& (overload_percent>= 25), overload_percent * 10, overload_percent)
    overload_percent = np.where(overload_percent>= 20, overload_percent * 20, overload_percent)

    total_penalty = np.sum(overload * overload_percent)
    return total_penalty

def calculate_fitness_storage_cost(stock):
    # штраф стоимость хранения- если общий остаток больше 40тонн и увеличивается
    # - 3000 за каждые полные/неполные 3 тонны + хранение каждого килограмма 1 р в день

    over_pen = []
    stor_pen = []
    overage_penalty_per_three_tons = 5000
    storage_penalty_per_kilogram = 1 # было 0.3 - увеличил плату до 1
    total_penalty = 0
    stock_limit=60000
    stock_max_limit=90000
    penalty_multiplier=5

    stock_a_day = stock.sum(axis=0)[1:]
    prev_date = formatted_dates[0]

    for i, date in enumerate(formatted_dates[1:], start=1):
        current_stock = stock_a_day[date]
        previous_stock = stock_a_day[prev_date]

        if previous_stock > stock_limit and current_stock > previous_stock:
            overage = current_stock - previous_stock
            overage_in_three_tons = math.floor(overage / 3000)
            total_penalty += overage_in_three_tons * overage_penalty_per_three_tons
        # плата storage_penalty_per_kilogram за каждый килограмм и каждый день хранения свыше stock_limit
        if current_stock > stock_limit:
            total_penalty += current_stock * storage_penalty_per_kilogram
        # двойная доп.плата storage_penalty_per_kilogram за каждый килограмм и каждый день хранения свыше stock_max_limit
        if current_stock>stock_max_limit:  # все что больше stock_max_limit - в penalty_multiplier раза дороже хранить
            total_penalty+=(current_stock-stock_max_limit) * storage_penalty_per_kilogram*penalty_multiplier

        prev_date = date
    return total_penalty


def calculate_fitness_order_completion(stock):
    # за каждый недопоставленные килограмм штраф 100 рублей
    penalty_per_kilogram = 100
    total_penalty = 0

    stock = np.array(stock[dates])
    total_penalty = abs(np.sum(stock[stock < 0] * penalty_per_kilogram))

    return total_penalty



def calculate_fitness_minstock_completion(stock, transit):
    # за отсутствие в дне минимального остатка штраф penalty_per_kilogram,
    # а за превышение троекратного минимального остатка 50 рублей

    # Логично смотреть переходящий остаток на последний день, а иначе
    # штрафуется за большой заказ позиции, которая заказывается редко

    # было 6 и 50 и 1,6
    penalty_per_kilogram = 4
    max_stock_penalty_multiplier = 30
    excessive_stock_coef=1.5

    ostatok_array = transit['transit_stock'].values
    stock_array = np.array(stock[dates])
    transit_stock_array = np.subtract(stock_array, ostatok_array[:, np.newaxis])

    # Вычисляем штрафы за каждый день за отсутствие а складе транзитного запаса
    daily_penalties = abs(np.where(transit_stock_array > 0, 0, transit_stock_array) * penalty_per_kilogram)

    # Проверяем условие для последнего дня, не должно быть остатка превышающего транзит в 2 раза
    last_day_stock = stock_array[:, -1]
    last_day_transit = ostatok_array[:]
    excessive_stock = last_day_stock - last_day_transit
    excessive_stock=np.where(excessive_stock>0,excessive_stock,0)
    if np.any(excessive_stock > excessive_stock_coef * ostatok_array):

        # Если есть день с превышающим остатком, применяем штраф
        mask = excessive_stock > excessive_stock_coef * ostatok_array
        excessive_stock_penalty = np.sum(
            (excessive_stock[mask] - ostatok_array[mask]) *
            max_stock_penalty_multiplier)
        total_penalty = np.sum(daily_penalties) + excessive_stock_penalty
    else:
        # В противном случае, применяем только штрафы за каждый день
        total_penalty = np.sum(daily_penalties)

    return total_penalty



def calculate_fitness_minimize_product_variety(production_plan):
    exponent = 2.2  # Показатель степени для полного 2.6
    no_change_days = args.no_change_days
    production_plan = np.array(production_plan[dates[no_change_days:]])
    # Считаем количество видов продукции в каждом дне,
    # штраф основание в степени кол-во видов выпускаемой продукции

    product_variety_per_day = np.where(production_plan > 0, 1, 0)
    product_variety_per_day = np.sum(product_variety_per_day, axis=0)
    penalty_per_day = exponent ** product_variety_per_day
    penalty = np.sum(penalty_per_day)

    return penalty


def total_production_cost(production_plan, stop_days):
    min_production = 7000
    low_prod_cost = 100000 #требует уточнения
    exponent=math.log(low_prod_cost)/math.log(min_production)
    stop_day_cost = 100000
    wheekend_cost = 30000
    total_cost = 0
    # если в день что-то делают и делают меньше чем min_production_cost = штраф low_prod_cost

    pp_array = np.array(production_plan)[:, 1::]
    pp_sum = np.sum(pp_array, axis=0)
    production_deficit = np.where(pp_sum < min_production, min_production-pp_sum, 0)
    working_day = np.where(pp_sum > 0, 1, 0)

    total_cost += np.sum((production_deficit**exponent) * working_day)

    #   если в день что-то делают и это выходной день - штраф за выходной

    dates_days_of_week = dates.apply(lambda x: get_day_type(x))
    dates_weekends = np.where(dates_days_of_week > 4, 1, 0)
    working_weekends = np.where(dates_weekends * pp_sum > 0, 1, 0)
    total_cost += np.sum(working_weekends * wheekend_cost)

    #   если в стоп-день работают то это штраф stop_day_cost

    stopdays_plan = np.where(dates.isin(stop_days), 1, 0)
    stop_days_work = np.sum(stopdays_plan * pp_sum)
    total_cost += stop_days_work * stop_day_cost

    return total_cost


def random_change_prod(production_plan, change=100, chance=0.5):
    prod = production_plan.copy()
    prod.iloc[:, 1:] = prod.iloc[:, 1:].applymap(
        lambda num: random_change(num, change, chance))
    return prod


# Функция для случайного изменения числа

def random_change(num, change, chance):
    if num == 0:
        if random.randint(0, 100) / 100 < chance / 2:
            num = random.randint(0, change)

    if random.randint(0, 100) / 100 < chance:
        num += random.randint(0, min(int(num), int(change)))

    return num


def random_generate(num, change=4000, chance=1 / 13):
    if random.randint(0, 100) / 100 < chance:
        num = random.randint(0, int(change))
    else:
        num = 0
    return num


def product_plan_generate(production_plan, change=100, chance=0.1):
    prod = production_plan.copy()
    prod[prod.columns[1:]] = prod.iloc[:, 1:].applymap(lambda num: random_generate(num))
    return prod


def calculate_fitness_production_changes(production_plan, production_plan_suspect):
    penalty_multiplier = 0.2
    nochange_penalty = 100000000
    # tomorrow_penalty = 100000
    no_change_days=args.no_change_days # сколько дней не менять. 1- текущий, 2 - текущий и завтра и т.д.

    # Преобразование датафреймов в массивы numpy
    plan_array = production_plan.iloc[:, 1:].to_numpy()
    suspect_array = production_plan_suspect.iloc[:, 1:].to_numpy()

    # Вычисление модуля разности массивов
    difference_array = np.abs(suspect_array - plan_array)
    #     штрафы за изменения в 1-2 дни
    # extra_penalty = (difference_array[:, 0] * nochange_penalty)  # штраф за изменения в текущем дне
    extra_penalty = (difference_array[:, :no_change_days] * nochange_penalty)  # штраф за изменения в текущем дне
    # Умножение на коэффициент штрафа
    penalty_array = difference_array * penalty_multiplier
    # если планируем в пятницу то не меняем пятницу, субботу и понедельник
    # если планируем в субботу то не меняем субботу, воскресенье и понедельник

    # Вычисление общего штрафа
    total_penalty = np.sum(penalty_array) + np.sum(extra_penalty)
    return total_penalty


# начнем оформлять фитнессфункции

def calculate_total_fitness(stock, orders, transit_stock, initial_production_plan, production_plan_suspect, wc_resource,
                            resource_need):
    fitness_order_completion = calculate_fitness_order_completion(stock)
    fitness_minimal_stock = calculate_fitness_minstock_completion(
        stock, transit_stock)
    fitness_storage_cost = calculate_fitness_storage_cost(stock)
    fitness_minimize_product_variety = calculate_fitness_minimize_product_variety(
        production_plan_suspect)
    fitness_minimize_equipment_overloads = calculate_fitness_equipment_overloads(
        wc_resource, resource_need)
    fitness_total_production_cost = total_production_cost(production_plan_suspect, stop_days)
    fitness_production_plan_change = calculate_fitness_production_changes(
        initial_production_plan, production_plan_suspect)
    # Дополните код для остальных критериев
    total_fitness = []
    total_fitness.append(fitness_order_completion)
    total_fitness.append(fitness_minimal_stock)
    total_fitness.append(fitness_storage_cost)
    total_fitness.append(fitness_minimize_product_variety)
    total_fitness.append(fitness_minimize_equipment_overloads)
    total_fitness.append(fitness_total_production_cost)
    total_fitness.append(fitness_production_plan_change)

    # Дополните код для остальных критериев
    total_fitness_sum = pd.Series(total_fitness).sum()
    total_fitness.append(total_fitness_sum)

    return total_fitness

def generate_random_prod_plan(all_nomenclature, dates, min=0, max=5000, step=200, chance=0.1):
    # Создаем пустой DataFrame с колонкой "Номенклатура" и колонками дат
    df = pd.DataFrame()
    df['Номенклатура'] = all_nomenclature
    df[dates] = 0
    # Создаем массив случайных значений с заданными параметрами
    size_x = len(all_nomenclature)
    size_y = len(dates)
    chance_num = 1 / chance
#     маска значний с вероятностью *chance*
    i_shape = np.random.randint(chance_num, size=(
        size_x, size_y)) // (chance_num - 1)

    individual = np.random.randint(max, size=(size_x, size_y))
    individual = individual // step * step
    individual = individual * i_shape

    # Заполняем значения в DataFrame
    df.iloc[:, 1:] = individual

    return df


# сортируем особи
def pop_sort(osobs, fitness_key):
    # print(len(osobs))
    osobs.sort(key=lambda x: x.fitness[fitness_key], reverse=False)
    return osobs


def roulette_selection(population, num_parents=2):
    #     создадим вероятность выбора каждой в зависимости от фитнесса
    #      чем ниже фитнесс тем веорятней выбор
    fitness_values = 1 / (np.array([item.fitness[7] for item in population])) * 1000000000
    std = np.std(fitness_values)
    fitness_values = fitness_values / std
    total_fitness = np.sum(fitness_values)
    probabilities = (fitness_values) / total_fitness

    selected_indices = np.random.choice(
        len(population), size=num_parents, p=probabilities, replace=False)
    selected_parents = [copy.deepcopy(population[idx]) for idx in selected_indices]
    if len(selected_parents) != num_parents:
        selected_parents = roulette_selection(population, num_parents)

    # если нужно для кроссовера - то не скрещивать одинаковых
    if num_parents==2:
        if selected_parents[0].production_plan.equals(selected_parents[1].production_plan):
            selected_parents = roulette_selection(population, num_parents)
    # если нужен только один экземпляр(например для мутации) то возвращать его
    if num_parents==1:
        selected_parents=selected_parents[0]

    return selected_parents

import numpy as np
import copy

def select_parents_on_num(population, num_parents=2):
    # Создаем вероятность выбора каждой особи в зависимости от ее положения в популяции
    probabilities = 1 / (np.arange(1, len(population) + 1))

    # Нормализуем вероятности
    probabilities /= np.sum(probabilities)

    # Выбираем индексы особей с учетом вероятностей
    selected_indices = np.random.choice(
        len(population), size=num_parents, p=probabilities, replace=False)

    # Проверяем, что выбранные индексы разные
    while len(set(selected_indices)) < num_parents:
        selected_indices = np.random.choice(
            len(population), size=num_parents, p=probabilities, replace=False)

    selected_parents = [copy.deepcopy(population[idx]) for idx in selected_indices]

    if len(selected_parents) != num_parents:
        selected_parents = select_parents_on_num(population, num_parents)

    # Исключаем особи с одинаковыми production_plan
    if selected_parents[0].production_plan.equals(selected_parents[1].production_plan):
        selected_parents = select_parents_on_num(population, num_parents)

    return selected_parents

def create_boolean_matrix(rows, cols, true_count):
    if true_count > rows * cols:
        print(f"Number of 'true' {true_count} values cannot exceed matrix size- {rows * cols}. I made it equal {rows * cols-1}")
        true_count=rows * cols-1


    matrix = np.full((rows, cols), False)
    flat_indices = np.random.choice(rows * cols, true_count, replace=False)
    row_indices, col_indices = np.unravel_index(flat_indices, (rows, cols))
    matrix[row_indices, col_indices] = True
    return matrix


def crossover(parent1, parent2, crossover_rate=0.2):
    global stock
    global orders
    rows = len(must_use_nomenclature)
    columns = len(dates)
    #     crossover_rate =

    df1 = pd.DataFrame(parent1.production_plan)
    df2 = pd.DataFrame(parent2.production_plan)

    # Создаем булевую маску и заменяем значения в столбце 0 на True

    mask = create_boolean_matrix(
        rows, columns + 1, math.floor(crossover_rate * rows * columns))
    mask[:, 0] = True

    # Выбираем значения из df1 и df2 в соответствии с маской
    df1_transit = copy.copy(df1)
    df1_transit_values = df1_transit.values

    df1_values = df1.values
    df1_values[mask] = df2.values[mask]

    df2_values = df2.values
    df2_values[mask] = df1_transit.values[mask]

    # Создаем новые DataFrame с обновленными значениями
    new_df1 = copy.copy(pd.DataFrame(df1_values, columns=df1.columns))
    new_df2 = copy.copy(pd.DataFrame(df2_values, columns=df2.columns))

    # Создаем новые экземпляры Individual с обновленными DataFrame
    child1 = Individual(new_df1, stock, orders)
    child2 = Individual(new_df2, stock, orders)

    return child1, child2


def mutation(osob, mutation_rate, mutation_chance):
    new_osob = copy.deepcopy(osob)  # Создаем копию особи
    df = new_osob.production_plan
    df_array1 = np.array(df)
    rows = len(must_use_nomenclature)
    columns = len(dates)

    #     мутация - изменение того, что есть
    matrix_change = np.array(
        [[random.randint(int((1 - mutation_rate) * 100), int((1 + mutation_rate) * 100)) for _ in range(columns + 1)]
         for _ in range(rows)])
    matrix_change = matrix_change / 100
    true_count = mutation_chance * rows * columns

    mask = create_boolean_matrix(
        rows, columns + 1, round(true_count))
    # заменил все mask[:,0]=False на ниже, чтобы лишний раз не менять мутацией запретные ячейки
    mask[:, 0:args.no_change_days+1] = False
    df_array1[mask] = df_array1[mask] * matrix_change[mask]

    # мутация добавление
    true_count = mutation_chance /2 * rows * columns
    # matrix=np.array()
    matrix = np.array([[random.randint(0, 4000)
                        for _ in range(columns + 1)] for _ in range(rows)])
    mask = create_boolean_matrix(
        rows, columns + 1, round(true_count))
    mask[:, 0:args.no_change_days+1] = False
    df_array1[mask] = matrix[mask]

    # мутация добавление нулей
    matrix_nul = np.zeros_like(df_array1)

    true_count = mutation_chance/2 * rows * columns
    mask = create_boolean_matrix(
        rows, columns + 1, round(true_count))
    mask[:, 0:args.no_change_days+1] = False
    df_array1[mask] = df_array1[mask] * matrix_nul[mask]
    #     оформление в особь
    columns = new_osob.production_plan.columns
    new_osob.production_plan = pd.DataFrame(df_array1)
    new_osob.production_plan.columns = columns
    new_osob.recalc()
    return new_osob



# Генерация плана
def generate_random_prod_plan(nomenclature, dates):
    df = pd.DataFrame()
    df['Номенклатура'] = nomenclature

    # просто случайно заполняет числами

    # Создание массива NumPy с уникальными случайными значениями от 0 до 4000 для каждой даты
    random_values = np.random.randint(1000, size=(len(nomenclature), len(dates)))

    # Создание маски с вероятностью 10%
    mask = np.random.choice([True,False],size=(len(nomenclature), len(dates)),p=[0.9,0.1])

    # Применение маски к массиву случайных значений
    random_values[mask] = 0
    # Присвоение значений массива DataFrame
    df[dates] = random_values
    # старый метод долгий слишком

    # medians = transit_stock['transit_stock']
    # stds = orders.iloc[:, 1:].max(axis=1)
    # for row in range(len(nomenclature)):
    #     median = medians[row]
    #     std = stds[row]
    #     for col in dates:
    #         value = generate_value(median, std)
    #         df.at[row, col] = value

    return df


def population_info(population, key=7):
    # print(len(population))
    # print(population[0].fitness)
    fitness = [individual.fitness[key] for individual in population]
    fitness = pd.Series(fitness)
    #     fitness = pd.Series(population)
    #     fitness = fitness.apply(lambda x: x.fitness[7])
    mean, count, min = fitness.describe()[['mean', 'count', 'min']]
    print(f'размер популяции = {count:,} особей, среднее - {mean:,.0f}, лучшее - {min:,.0f}')


def drop_duplicates(population):
    set_osob = []
    set_osob_fitness = []
    for name in population:
        if name.fitness not in set_osob_fitness:
            set_osob.append(name)
            set_osob_fitness.append(name.fitness)
        else:
            pass
    return set_osob


def save_population(population, filename):
    # дополнить флаг args.low_net который реже созраняет
    max_retries = 60  # Максимальное количество попыток (10 минут)
    retry_delay = 5+random.randint(1,20)  # Задержка между попытками (в секундах)
    retry_count = 0  # Счётчик ошибок
    while retry_count < max_retries:
        try:
            lock = multiprocessing.Lock()
            with lock:
                with gzip.open(f'{filename}.gz', 'wb') as file:
                    pickle.dump(population, file)
                break
        except (Exception) as e:
            # function_name = traceback.extract_stack(None, 2)[0][2]
            current_function_name = inspect.currentframe().f_code.co_name
            print(f'ошибка1 в {current_function_name} : {str(e)}')
            time.sleep(retry_delay)
            retry_count += 1
        # finally:
        #     lock.release()


# Загрузка population из файла
def load_population(filename):
    lock = multiprocessing.Lock()
    uncompressed_data=[]
    loaded_population=[]
    with lock:
        try:
            with gzip.open(f'{filename}.gz', 'rb') as file:
            # Разархивируем данные
                uncompressed_data = file.read()
            loaded_population.extend(pickle.loads(uncompressed_data))

        except (Exception) as e:
            current_function_name = inspect.currentframe().f_code.co_name
            # print(f'ошибка2 в {current_function_name} : {str(e)}')
            # print(e)
            pass
        # finally:
        #     lock.release()

    return loaded_population


def new_population(count):  # создает случайную популяцию из count особей
    population = []
    population_size = count
    for _ in range(population_size):
        suspect_plan = generate_random_prod_plan(must_use_nomenclature, dates)
        osob = Individual(suspect_plan, stock, orders)  # Здесь нужно написать функцию для генерации случайного индивида
        population.append(osob)
    return population


# в зависимости от базы и поколения возвращает кол-во особей к генерации
def calculate_population_size(current_generation, add_new_base):
    new_individuals = math.ceil(add_new_base * (1.2) ** (-0.6 * current_generation))
    return max(new_individuals // 1, 5)

def local_search(population, mutation_rate=0.05, mutation_chance=0.2, mutation_num=10):
    # Применение локального посика
    new_mutants = []
    for individual in population:
        for _ in range(mutation_num):
            new_mutant = mutation(individual, mutation_rate=mutation_rate, mutation_chance=mutation_chance)
            new_mutants.append(new_mutant)
    return new_mutants


def load_populations_and_merge(population_file_pattern, current_pid,population):
    # загружает и удаляет сохранения из других процессов
    uncompressed_data = []
    loaded_population=[]
    counter=0
    try:
        for filename in os.listdir(args.work_dir):
        # for filename in os.listdir(args.work_dir):
            # прерывать загрузку если больше 5 успешных загрузок
            if counter>5:
                return loaded_population
            if filename.startswith(population_file_pattern) and filename.endswith('.pkl.gz'):
                try:
                    # Извлекаем значение pid из имени файла
                    file_pid = ((filename.split('_pid'))[1].split('.pkl')[0])

                    # Проверяем, что это не текущий evo_line
                    if file_pid != current_pid:
                        filename = os.path.join(args.work_dir, filename)
                        lock = multiprocessing.Lock()
                        with lock:
                            with gzip.open(f'{filename}', 'rb') as file:
                                # Разархивируем данные
                                uncompressed_data = file.read()
                        loaded_population.extend(pickle.loads(uncompressed_data))
                        counter+=1
                        try:
                            if population[0].fitness[7]<=loaded_population[0].fitness[7]:
                                os.remove(f'{filename}')
                        except (Exception) as e:
                            current_function_name = inspect.currentframe().f_code.co_name
                            pass
                        # finally:
                        #     lock.release()
                except (Exception) as e:
                    current_function_name = inspect.currentframe().f_code.co_name
                    pass
    except (Exception) as e:
        current_function_name = inspect.currentframe().f_code.co_name
    return loaded_population

def load_all_pickles(population_file_pattern):
    loaded_population=[]
    for filename in os.listdir(args.work_dir):
        if filename.startswith(population_file_pattern) and filename.endswith('.pkl'):
            try:
                filename = os.path.join(args.work_dir, filename)
                lock = multiprocessing.Lock()
                with lock:
                    with open(filename, 'rb') as file:
                        loaded_population.extend(pickle.load(file))
                    os.remove(f'{filename}')
            except (Exception) as e:
                current_function_name = inspect.currentframe().f_code.co_name
                # print(f'ошибка5 в {current_function_name} : {str(e)}')
                # function_name = traceback.extract_stack(None, 2)[0][2]
                # print(f'ошибка в {function_name} : {str(e)} with {filename}')
                pass
            # finally:
            #     lock.release()
    return loaded_population

def evolution_v1(population, evo_line, add_new_size, pid, end_time): # альтернатива формированию нового поколения
    global max_generations
    global stock
    global orders

    mytable.sortby = 'evo'

    # non_productive = 10
    prev_best_fitness = float('inf')
    key = 7
    # max_population_size = 100  # используется для обрезки популяции
    loaded_population = []
    current_pid = gethostname() + str(pid)
    population_file_name = f"{population_file_pattern}_pid{current_pid}.pkl"
    # load_param = 100
    num_multiplier=1
    start_time = time.time()
    last_time_send=time.time()
    full_path = os.path.join(args.work_dir, population_file_name)
    stop_count=0
    next_population=[]
    fitness_history=[]
    generation_counter=0

    if args.no_exchange:
        # при наличии флага no_exchange все процессы работают независимо друг от друга в течение значения которое флагу присвоено
        # после этого начинаются обмены между разными процессами и в течение extra_crossover_generations
        # кол-во кроссоверов увеличено до extra_crossover_num
        extra_crossover_generations = 10
        extra_crossover_num = 2000
        exchange_start_time=start_time+args.no_exchange*60
        print (f'start_time {start_time}, exchange_start_time {exchange_start_time}, diff is {args.no_exchange*60}')
    else:
        extra_crossover_generations = 0
        extra_crossover_num = 0
        exchange_start_time = start_time

    for generation in range(100000):
        try:


            evolution_start = time.time()

            # загружаем лучшие особи из других эволюций
            if time.time()>exchange_start_time:
                generation_counter+=1
                loaded_population = load_populations_and_merge(population_file_pattern, current_pid,population)

                # if (len(loaded_population)>0) and (generation < args.no_exchange+10):
                #     for osob in loaded_population:
                #         osob.recalc()
                population.extend(loaded_population)
                loaded_population = []

            if generation<1000:
                mutation_rate = random.randint(1, 20) / 100
                mutation_chance = random.randint(1, 10) / 100
            else:
                mutation_rate = random.randint(1, 10 ) / 100
                mutation_chance = random.randint(1, 6 ) / 100




            # заменил определение кол-ва мутаций - 10% популяции
            # mutation_num = math.ceil(len(population) * 0.2) + 10+generation//30
            mutation_num=20
            crossover_num=40
            if (time.time() < exchange_start_time):
                crossover_num=100
                mutation_num=100
                mutation_rate=random.randint(1, 100) / 100
                mutation_chance = random.randint(1, 6) / 100
            # заменил в определении mutation_num и crossover_num минимальное кол-во на 10, было 50
            crossover_rate = random.randint(10, 90) / 100
            # crossover_num = 10 + math.floor(2.7 ** (0.6 * random.randint(1, 7)))*num_multiplier
            # crossover_num=len(population)*0.7//1+1
            # если выбрано значение no_exchange(не обмениваться осоями между генерациями в течение скольки-то поколений
            # то после того, как обмен между поколениями начался - увеличить кол-во  кроссоверов на какое-то кол-во поколений
            # кол-во поколений до intense_crossover_generations, кол-во кроссоверов + extra_crossover


            if (args.no_exchange>0)&(time.time()>exchange_start_time)&(extra_crossover_generations>generation_counter):
                crossover_num+=extra_crossover_num

            if evo_line==0:
                if random.randint(0,100)>90:
                    mutation_num*=2
                    crossover_num*=2
                    # mutation_chance*=2
                    # mutation_rate*=2

            if stop_count>0:
                mutation_num+=int(np.log(stop_count)*20)
                crossover_num+=int(np.log(stop_count)*40)




            # добавим в стартовую популяцию несколько случайных особей
            if generation <100:
                population_size = calculate_population_size(generation, add_new_base=add_new_size)
                population.extend(new_population(population_size))

            if len(population)>0:
                population = pop_sort(population, key)

            # кроссоверы
            for _ in range(int(crossover_num)):
                parent1, parent2 = select_parents_on_num(population, num_parents=2)
                child1, child2 = crossover(
                    parent1, parent2, crossover_rate=crossover_rate)

                next_population.extend([child1, child2])

            # Применение мутаций
            for _ in range(int(mutation_num)):
                mutant_candidate=roulette_selection(next_population,
                                                      num_parents=1)
                new_mutant = mutation(mutant_candidate, mutation_rate=mutation_rate, mutation_chance=mutation_chance)
                next_population.append(new_mutant)



            # Добавляем в новую популяцию элиту из исходной
            next_population.extend(population[:2])
            # сортируем новую популяцию по фитнесу
            if len(next_population)>0:
                next_population = pop_sort(next_population, key)

            # удалим дубликаты
            next_population = drop_duplicates(next_population)

            best_fitness = next_population[0].fitness[7]

            if prev_best_fitness > best_fitness:
                prev_best_fitness = best_fitness
                stop_count = 0
            else:
                stop_count += 1

            # телеграм только 0 поток и только в мастерской вариации и только раз в 100 поколений и только если телеграм заказывали
            if (args.master) and (evo_line == 0) and ((generation % 100) == 0) and (args.telegram):
                # fitness_performance=(best_fitness-last_time_fitness)/(time.time()-last_time_send)*3600
                send_message(
                    f'Минимальная стоимость:{best_fitness//1}, {generation} попыток, работает {(time.time() - start_time) // 3600} часов из {int((end_time - start_time) / 3600)} часов. Уровень недопоставок - {next_population[0].fitness[0]//100} кг')
                last_time_send=time.time()
                last_time_fitness=best_fitness

            if (evo_line == 0) and ((generation % 100) == 0):
                with gzip.open(f'fitness_history{pid}.gz', 'wb') as file:
                    pickle.dump(fitness_history, file)


            # if (time.time() > end_time) or (stop_count>200):
            if (time.time() > end_time):
                mytable.clear_rows()
                mytable.add_row(
                    [evo_line, generation, len(population), stop_count, mutation_rate, mutation_chance, crossover_rate,
                     mutation_num, crossover_num, int(next_population[0].fitness[7]), int(next_population[0].fitness[0]),
                     int(next_population[0].fitness[1]), int(next_population[0].fitness[2]), int(next_population[0].fitness[3]),
                     int(next_population[0].fitness[4]), int(next_population[0].fitness[5]), int(next_population[0].fitness[6]),
                     time.strftime("%H:%M:%S", time.gmtime(time.time() - evolution_start))])
                print(mytable)

                save_population(next_population, full_path)
                return next_population

            # if generation > args.no_exchange:
            if len(next_population)>3:
                save_pop=next_population[:1]
            else:
                save_pop=next_population
            save_population(save_pop, full_path)

            mytable.clear_rows()
            mytable.add_row([evo_line, generation, len(next_population),stop_count, mutation_rate, mutation_chance, crossover_rate,
                             mutation_num, crossover_num, int(next_population[0].fitness[7]), int(next_population[0].fitness[0]),
                             int(next_population[0].fitness[1]), int(next_population[0].fitness[2]), int(next_population[0].fitness[3]),
                             int(next_population[0].fitness[4]), int(next_population[0].fitness[5]), int(next_population[0].fitness[6]),
                             time.strftime("%H:%M:%S", time.gmtime(time.time() - evolution_start))])
            print(mytable)

            # sharp_size=min(len(next_population),max_population_size)
            #
            # if stop_count>0:
            #     sharp_size+=int(np.log(stop_count)*40)

            population=next_population

            if evo_line==0:
                fitness_history.append([generation, next_population[0].fitness])
            next_population = []
        except (Exception) as e:
            traceback.print_exc()
            new_pop = []
            current_function_name = inspect.currentframe().f_code.co_name
            print(f'ошибка7 в {current_function_name} : {str(e)}')
            pass
    return next_population

def set_options(pop_size=30, max_pop_size=100, max_gen=100, m_c=0.5, m_r=0.5, cr_r=0.5):
    global population_size  # используется для добавления новых особей в популяцию
    population_size = pop_size
    global max_population_size  # используется для обрезки популяции
    max_population_size = max_pop_size
    global max_generations  # кол-во поколений для процесса
    max_generations = max_gen
    global mutation_chance  # option not involved in process
    mutation_chance = m_c
    global mutation_rate  # option not involved in process
    mutation_rate = m_r
    global crossover_rate  # option not involved in process
    crossover_rate = cr_r


def calculate_seconds_to_next_hour(hour):
    # Получаем текущее время
    now = datetime.now()

    # Создаем объект datetime для ближайшего 7:00:00
    next_hour = datetime(now.year, now.month, now.day, hour, 0, 0)
    if now.hour >= hour:
        # Если текущее время уже прошло указанный час, добавляем 1 день
        next_hour += timedelta(days=1)

    # Вычисляем разницу в секундах до ближайшего 7:00:00
    time_difference = next_hour - now
    seconds_to_next_hour = time_difference.total_seconds()
    return seconds_to_next_hour


def evolve(x, count, seed, end_time):
    set_options(pop_size=count, max_pop_size=200, max_gen=3000)
    random.seed(seed)  # Устанавливаем уникальное зерно для генератора случайных чисел в каждом процессе
    population = new_population(count)
    # time.sleep(random.randint(0,20))
    # доавим оригинальную особь в каждую популяцию
    population.append(initial_osob)
    # for osob in population:
    #     osob.recalc()
    add_new_size=count
    pid = getpid()
    try:
        new_pop=evolution_v1(population, x, population_size,pid,end_time)
    except (Exception) as e:

        traceback.print_exc()
        new_pop=[]
        current_function_name = inspect.currentframe().f_code.co_name
        print(f'ошибка7 в {current_function_name} : {str(e)}')
        pass

    return new_pop


# Функция локального поиска перебирает все параметры плана производства изменяя его на +- delta
# потом считает для каждого фитнесс функцию
# те изменения которые привели к уменшению функии повторяются count раз
# после чего лучшая особь возвращается

import copy

import random
from datetime import datetime

def local_optimization(individual, delta, count, dates):
    individual_copy = copy.deepcopy(individual)
    best_individual = copy.deepcopy(individual)

    day_type = get_day_type(dates[0])
    if day_type == 4:
        no_estimate_days = 4
    elif day_type == 5:
        no_estimate_days = 3
    else:
        no_estimate_days = 2

    for _ in range(count):
        # Случайный выбор ячейки в таблице
        day = random.randint(no_estimate_days + 1, len(individual.production_plan.columns) - 1)
        nomenclature = random.choice(individual.production_plan.index)
        parameters_changed = True
        while parameters_changed:
            parameters_changed = False
            original_value = best_individual.production_plan.iat[nomenclature, day]

            # Увеличение параметра на delta
            best_individual.production_plan.iat[nomenclature, day] += delta
            best_individual.recalc()  # Пересчет остатков и фитнеса

            if best_individual.fitness[7] < individual_copy.fitness[7]:
                individual_copy.production_plan.iat[nomenclature, day] = best_individual.production_plan.iat[nomenclature, day]
                individual_copy.recalc()  # Пересчет остатков и фитнеса
                parameters_changed = True
            else:
                # Уменьшение параметра на delta
                best_individual.production_plan.iat[nomenclature, day] -= 2 * delta
                if best_individual.production_plan.iat[nomenclature, day] >= 0:
                    best_individual.recalc()  # Пересчет остатков и фитнеса
                else:
                    best_individual.production_plan.iat[nomenclature, day] = 0
                    best_individual.recalc()  # Пересчет остатков и фитнеса

                if best_individual.fitness[7] < individual_copy.fitness[7]:
                    individual_copy.production_plan.iat[nomenclature, day] = best_individual.production_plan.iat[nomenclature, day]
                    individual_copy.recalc()  # Пересчет остатков и фитнеса
                    parameters_changed = True
                else:
                    # Восстановление оригинального значения параметра
                    best_individual.production_plan.iat[nomenclature, day] = original_value
                    best_individual.recalc()  # Пересчет остатков и фитнеса

    return individual_copy


# мы создадит класс индивид
# в нем будут методы обращения к плану производства и формированию стоков и потребности в ресурсах.
# создадим класс особь - Individual

class Individual:
    def __init__(self, production_plan_suspect,stock, orders):
        self.production_plan = production_plan_suspect

        self.stock = copy.copy(stock)
        self.fill_stock(orders)
        #         self.resource_need = [123]
        self.fill_resource_need()
        self.fitness = None
        self.count_fitness()

    def recalc(self):
        #         self.production_plan = production_plan
        self.fill_stock(orders)
        self.fill_resource_need()
        self.count_fitness()

    #  нужно векторизировать
    def fill_stock(self, orders):

        for date_i in range(0, len(dates) - 1):
            # Получаем текущие заказы, остатки и планы производства для данной даты
            current_orders = np.array(orders[dates[date_i]].tolist())
            current_stock = np.array(self.stock[dates[date_i]].tolist())
            # Проверяем, есть ли прошлый план производства для данной даты
            days_per_freez =1 # дублируем код count_stock, переменная сколько дней на заморозку.
            if date_i >= days_per_freez:
                current_plan = np.array(
                    self.production_plan[dates[date_i - days_per_freez]].tolist())
            # Если прошлого плана нет, заполняем его нулями
            else:
                current_plan = np.array([0] * len(current_orders))

            current_stock = np.where(current_stock < 0, 0, current_stock)

            # Вычисляем новые остатки на следующую дату
            stock_i = current_stock - current_orders + current_plan

            # Обновляем остатки на следующую дату
            self.stock[dates[date_i + 1]] = stock_i

    def fill_resource_need(self):
        prod_plan = self.production_plan  # Загружаем DataFrame с планами производства
        # Загружаем DataFrame с требованиями к рабочим центрам
        wc_req = dfs['wc_req']
        all_wc_needs = []  # Создаем пустой список для хранения всех wc_needs
        # Перебираем ключи и значения словаря type_to_names
        for prod_type, nomenclatures in type_to_names.items():
            if prod_type != 0:
                # Получаем коэффициенты использования рабочих центров для данного prod_type
                wc_coefficients = wc_req[prod_type]
                wc_coefficients = np.array(wc_coefficients)

                # Создаем Series для индексации DataFrame prod_plan
                nomenclature_filter = prod_plan['Номенклатура'].isin(
                    nomenclatures)

                # Суммируем планы производства для номенклатур данного типа
                plan_sum = prod_plan[nomenclature_filter].sum(axis=0)
                plan_sum = np.array(plan_sum[1:])

                # Рассчитываем потребности в ресурсах на каждый день
                wc_needs = np.outer(plan_sum, wc_coefficients).T
                all_wc_needs.append(wc_needs)

        # Суммируем все рассчитанные значения для получения итоговой потребности в ресурсах
        result_array = np.zeros_like(all_wc_needs[0])
        for array in all_wc_needs:
            result_array += array

        # Создаем DataFrame для хранения итоговой потребности в ресурсах
        raws = wc_req['Рабочий центр']
        columns = prod_plan.columns[1:]
        # Создаем атрибут resource_need внутри self и сохраняем туда DataFrame с вычисленными потребностями

        self.resource_need = pd.DataFrame(result_array, columns=columns)
        # вставим в начало колонку с рабочими центрами
        self.resource_need.insert(0, 'Рабочий центр', raws.values)

    def count_fitness(self):  # возвращает кортеж с фитнесами, [7] - итоговый
        self.fitness = calculate_total_fitness(
            self.stock, orders, transit_stock, dfs['production_plan'], self.production_plan, wc_resource,
            self.resource_need)




# задаем список дней когда планируем не работать
# dd-mm-YYYY
stop_days = ['07-06-2024','08-06-2024','09-06-2024','10-06-2024','11-06-2024']
holidays = ['01-01-2024', '02-01-2024', '03-01-2024', '04-01-2024', '05-01-2024', '06-01-2024',
            '07-01-2024', '08-01-2024', '23-02-2024', '08-03-2024','09-05-2024', '01-05-2024','01-09-2024', ]
stop_days.extend(holidays)

file_path=f'{args.work_dir}\исходные данные для планирования.xlsx'

dfs = load_excel(file_path)
dict_to_globals(dfs)

# устанавливаем формат даты для преобразований из заголовкой
date_format = '%d.%m.%Y %H:%M:%S'
# date_format = '%d.%m.%Y'


# находим саммую раннюю дату
first_date_orders = datetime.strptime(dfs['orders'].columns[2], date_format)
first_date_prod=datetime.strptime(dfs['production_plan'].columns[2], date_format)
# действительно сомнительно!!!!
# start_date=datetime.today() # сомнительна необходимость использования этой переменной. для анализа??
# start_date=min(start_date,first_date_prod,first_date_orders)
start_date=min(first_date_prod,first_date_orders)


# находим последнюю дату
num_columns_orders=len(dfs['orders'].columns)
num_columns_prod=len(dfs['production_plan'].columns)
# было бы неплохо сделать проверку, что лист wc_resource имеет ту же размерность что и потенциальны prod_plan

last_date_order=datetime.strptime(dfs['orders'].columns[num_columns_orders-1], date_format)+timedelta(days=1)
last_date_production_plan=datetime.strptime(dfs['production_plan'].columns[num_columns_prod-1], date_format)+timedelta(days=2)
last_date=max(last_date_order,last_date_production_plan)

resources = wc_req.iloc[:,0]

dfs['resource_need'] = initialize_resource_need(resources, start_date, last_date)

# Цикл по датафреймам
for key, df in dfs.items():

    # Цикл по колонкам
    for col in df.columns:
        # Попытка отформатировать как дату
        formatted = format_date(col)
        df.rename(columns={col: formatted}, inplace=True)

        # Сохранение датафрейма
        dfs[key] = df

# заменим в стоках колонку со словами на первую дату
dfs['stock'].rename(columns={'Количество (в ед. отчетов)': start_date.strftime('%d-%m-%Y')}, inplace=True)

# Добавляем инициализацию

desired_format = '%d-%m-%Y'
date_columns = create_date_columns(start_date, last_date)
formatted_dates = date_columns

# Добавление дат для каждого датафрейма
for key in ['orders','production_plan','stock', 'wc_resource']:
    df = dfs[key]
    dfs[key] = add_missing_dates(df, formatted_dates)

# переименуем первый столбец в wс_req в Рабочий центр
dfs['wc_req'].rename(columns={wc_req.columns[0]: 'Рабочий центр'}, inplace=True)

# сортирвка датафреймов по именам рабочих центров
dfs['wc_req']=dfs['wc_req'].sort_values(by='Рабочий центр',ignore_index=True)
dfs['wc_resource']=dfs['wc_resource'].sort_values(by='Рабочий центр',ignore_index=True)

all_nomenclature = set()
used_nomenclature=set()
# получим множество всей номенклатуры в списках с номенклатурой
for key in ['orders', 'stock', 'production_plan', 'week_history']:
    all_nomenclature.update(list(dfs[key]['Номенклатура']))

must_use_nomenclature=set()
must_use_nomenclature=set(dfs['production_plan'][dfs['production_plan'][date_columns[0:args.no_change_days]].gt(0).any(axis=1)]['Номенклатура'])
must_use_nomenclature.update(dfs['orders'][dfs['orders'][date_columns].gt(0).any(axis=1)]['Номенклатура'])


# Создаем словарь для хранения соответствия номенклатуры и вида производства
nomenclature_to_production = {}

nomenclature_to_production = fill_product_dict(dfs['orders'], all_nomenclature)
nomenclature_to_production.update(fill_product_dict(dfs['stock'], all_nomenclature))
nomenclature_to_production.update(fill_product_dict(dfs['production_plan'], all_nomenclature))
nomenclature_to_production.update(fill_product_dict(dfs['week_history'], all_nomenclature))

def remove_nonused_nomenclature(df,used_nomenclature):
    # Проверка наличия колонки "номернклатура" в DataFrame
    if 'Номенклатура' not in df.columns:
        raise ValueError("Колонка 'номенклатура' не найдена в DataFrame.")
    # Фильтрация DataFrame по нужной номенклатуре
    filtered_dataframe = df[df['номернклатура'].isin(used_nomenclature)]
    return filtered_dataframe


# Цикл по датафреймам
for sheet in ['stock', 'orders', 'production_plan','week_history']:
    # Добавление недостающих строк в датафрейм
    dfs[sheet] = add_missing_rows(dfs[sheet], all_nomenclature).fillna(0)
    # Попытка удалить колонку "Вид производства"
    try:
        dfs[sheet].drop(columns='Вид производства', inplace=True)
    except:
        pass

# очищаем датафреймы от паразитного 0
all_nomenclature.discard(0)

for df in dfs:
    # print(df)
    try:
        dfs[df] = dfs[df][dfs[df]['Номенклатура'] != 0]
        dfs[df].drop(columns='Вид производимой продукции',inplace=True)
    except:
        pass

all_nomenclature=pd.Series(list(all_nomenclature))
all_nomenclature.sort_values(inplace=True)

must_use_nomenclature=pd.Series(list(must_use_nomenclature))
must_use_nomenclature.sort_values(inplace=True)

# сортировака номенклатуры
for df in dfs:

    try:
        dfs[df].sort_values(by='Номенклатура', inplace=True)
        dfs[df].reset_index(drop=True, inplace=True)
        # print(f'в датафрейме {df} Номенклатура отсортирована')
    except:
        # print(f'в датафрейме {df} нет колонки Номенклатура')
        pass

dates=pd.Series(formatted_dates) # все даты какие есть

dfs['stock']=count_stock(dfs['stock'], dfs['orders'], dfs['production_plan'], dates)

# создаем словарь с видами производства, а значение - список номенклатуры
type_to_names={}

for key,value in nomenclature_to_production.items():
    if value not in type_to_names:
        type_to_names[value]=[key]
    else:
        type_to_names[value].append(key)

prod_plan = dfs['production_plan']  # Загружаем DataFrame с планами производства
wc_req = dfs['wc_req']  # Загружаем DataFrame с требованиями к рабочим центрам
all_wc_needs = []  # Создаем пустой список для хранения всех wc_needs

# Перебираем ключи и значения словаря type_to_names
for prod_type, nomenclatures in type_to_names.items():
    if prod_type != 0:
        # Получаем коэффициенты использования рабочих центров для данного prod_type
        wc_coefficients = wc_req[prod_type]
        wc_coefficients = np.array(wc_coefficients)

        # Создаем Series для индексации DataFrame prod_plan
        nomenclature_filter = prod_plan['Номенклатура'].isin(nomenclatures)

        # Суммируем планы производства для номенклатур данного типа
        plan_sum = prod_plan[nomenclature_filter].sum(axis=0)
        plan_sum = np.array(plan_sum[1:])

        # Рассчитываем потребности в ресурсах на каждый день
        # print(plan_sum)
        # print(wc_coefficients)
        wc_needs = np.outer(plan_sum, wc_coefficients).T
        all_wc_needs.append(wc_needs)

# Суммируем все рассчитанные значения для получения итоговой потребности в ресурсах
result_array = np.zeros_like(all_wc_needs[0])
for array in all_wc_needs:
    result_array += array

# Создаем DataFrame для хранения итоговой потребности в ресурсах
raws = wc_req['Рабочий центр']
columns = prod_plan.columns[1:]
# создадим датафрейм и добавим в него массив с вычисленными потребностями
dfs['resource_need'] = pd.DataFrame(result_array, columns=columns)
dfs['resource_need'].insert(0, 'Рабочий центр', raws.values)  # вставим в начало колонку с рабочими центрами

transit_stock = ABC_XYZ(week_history)
transit_stock = transit_stock[['Номенклатура', 'Вид производимой продукции', 'transit_stock']]
transit_stock=(add_missing_rows(transit_stock, all_nomenclature))
transit_stock.sort_values(by='Номенклатура',inplace=True)
transit_stock.fillna(0,inplace=True)

# Имя файла Excel
output_file = 'transit_stock.xlsx'
try:
    # Сохраняем DataFrame в Excel-файл
    transit_stock.to_excel(output_file, index=False)
except:
    # Проверяем, существует ли уже файл
    version = 1
    while os.path.exists(output_file):
        # Генерируем новое имя файла
        output_file = f'transit_stock_v{version}{extension}'
        version += 1

# Сохраняем DataFrame в Excel-файл
transit_stock.to_excel(output_file, index=False)

# transit_stock.to_excel('transit_stock.xlsx', index=False)
print('ok')

dfs['transit_stock']=transit_stock


orders=dfs['orders']
stock=dfs['stock']
production_plan=dfs['production_plan']
wc_resource=dfs['wc_resource']
wc_req=dfs['wc_req']
transit_stock=dfs['transit_stock']
resource_need=dfs['resource_need']

# не помню зачем этот код - сохранил, вдруг понадобиться
# new_plan = product_plan_generate(dfs['production_plan'])
# production_plan_suspect = random_change_prod(
#     production_plan, change=1000, chance=0.1)


# type_to_raw - словарь в котором ключ - тип продукции, а возвращает списко индексов
# Сначала сортируем серию, чтобы ее индексы были как в датафреймах
all_nomenclature.reset_index(drop=True, inplace=True)

# Создаем словарь, в котором ключ - тип, значение - список индексов
# такое ощущение, что это бесполезный словарь
type_to_raw = {}

for typ in type_to_names.keys():
    # Получаем индексы номенклатур данного типа
    list_index = list(all_nomenclature.loc[all_nomenclature.isin(type_to_names[typ])].index)
    # Сохраняем список индексов в словаре
    type_to_raw[typ] = list_index
# type_to_raw - словарьв котором ключ - тип продукции, а возвращает списко индексов

# Генерируем случайные данные с заданными параметрами
# generate_random_prod_plan(all_nomenclature, dates, step=100, min=0, max=4000)

# объявляем переменные для генетического алгоритма массивы данных

stock=dfs['stock'][dfs['stock']['Номенклатура'].isin(must_use_nomenclature)]
orders=dfs['orders'][dfs['orders']['Номенклатура'].isin(must_use_nomenclature)]
production_plan=dfs['production_plan'][dfs['production_plan']['Номенклатура'].isin(must_use_nomenclature)]
transit_stock=dfs['transit_stock'][dfs['transit_stock']['Номенклатура'].isin(must_use_nomenclature)]
wc_resource=dfs['wc_resource']
dfs['production_plan']=dfs['production_plan'][dfs['production_plan']['Номенклатура'].isin(must_use_nomenclature)]
initial_osob=Individual(production_plan, stock, orders)

fitness_history=[]

# заготовка заголовков столбцов для fitness_history, если преобразовывать его в pd.DataFrame
columns = ['выполнение', 'миностаток', 'хранение','разнообразие', 'перегруз', 'смены', 'изменения', 'итого']

fd = dates[0]
ld = dates[-1:].values[0]

# задает паттерн из дат планирования для файлов сохранения
population_file_pattern = f'plan_py for  {fd} - {ld}'

populations = []
# вывод на экран результатов эволюции
mytable=PrettyTable()
mytable.field_names=['evo', 'gen', 'pop_size', 'no_imp', 'mut_r', 'mut_c', 'cross_r','mut_n',
                     'cros_n', 'TFit', 'f_ords','min_st','st_cst', 'f_var',
                     'f_wc_over', 'prod','chang' ,'time']

if __name__ == '__main__':
    args=main()
    pc_name=gethostname()
    num_processes=get_cpunum(args.no_cpucount)
    # logging.basicConfig(filename="error_log.txt", level=logging.ERROR,
    #                     format="%(asctime)s [%(levelname)s]: %(message)s")

    # приоритет имеет время остановки, а не время работы
    if args.hours_to_work:
        if args.master:
            end_time = time.time() + args.hours_to_work * 3600  # Преобразовать часы в секунды и добавить к текущему времени
        else:
            end_time = time.time() + (args.hours_to_work-1) * 3600

    if args.end_time:
        if args.master:
            end_time=time.time()+calculate_seconds_to_next_hour(args.end_time)
        else:
            end_time = time.time() + calculate_seconds_to_next_hour(args.end_time)-3600

    send_message(f'{pc_name} подключился к планированию {fd}-{ld} в качестве {master_slave(args.master)}, {num_processes} потоков, на {round((end_time-time.time())//3600)} часов')

    with multiprocessing.Pool(num_processes) as pool:
        # Генерируем аргументы для пула процессов
        if args.no_exchange:
            count_number=100
        else:
            count_number=10
        arguments = [(x, random.randint(5, count_number), random.randint(1, 1000), end_time) for x in (range(num_processes))]
        print(np.around(arguments, decimals=0))
        # Запускаем эволюцию в нескольких процессах параллельно
        populations = pool.starmap(evolve, arguments)
        send_message(f'{pc_name} завершил планирование {fd}-{ld} в качестве {master_slave(args.master)}, {num_processes} потоков')

    # Объединяем результаты эволюции в одну популяцию
    print('pool is closed')
    hyper_pop = []

    for pop in populations:
        # print(pop[0].fitness[7])
        if len(pop)>0:
            hyper_pop.extend(pop)

    hyper_pop=pop_sort(hyper_pop,7)
    send_message(f'{pc_name}, лучший результат {hyper_pop[0].fitness[7]}')

    filename=f'{population_file_pattern}_{pc_name}_final.pkl'
    full_path = os.path.join(args.work_dir, filename)
    save_population(hyper_pop, full_path)

    # сделать итоги планирования в отдельном сообщении
    prod_plan=hyper_pop[0].production_plan
    stock_plan=hyper_pop[0].stock
    resource_plan=hyper_pop[0].resource_need
    types_of_product=pd.Series(must_use_nomenclature.apply(lambda x: nomenclature_to_production[x])).values
    prod_plan.insert(1, 'Вид производимой продукции', types_of_product)
    stock_plan.insert(1, 'Вид производимой продукции', types_of_product)
    # prod_plan.to_excel(f'best_plan_{pc_name}.xlsx', index=False)
    # stock_plan.to_excel(f'best_stock_{pc_name}.xlsx', index=False)
    # resource_plan.to_excel(f'best_resource_{pc_name}.xlsx', index=False)

    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill, Font
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation


    def autofit_columns(df, sheet, max_columns=None, max_rows=None):
        if max_columns is None:
            max_columns = len(df.columns)
        if max_rows is None:
            max_rows = len(df)

        for column_index, (column_name, column_data) in enumerate(df.iloc[:, :max_columns].items(), 1):
            max_length = 0
            for row_index, cell in enumerate(column_data.astype(str), 1):
                try:
                    if len(cell) > max_length:
                        max_length = len(cell)
                    # Устанавливаем перенос текста по строкам для самых широких колонок
                    if column_index in [1, 2]:  # Укажите индексы колонок, которые должны быть многострочными
                        sheet.cell(row=row_index, column=column_index).alignment = Alignment(wrapText=True)
                        adjusted_width = 40  # Установите желаемую ширину для двухстрочного текста
                    else:
                        adjusted_width = (max_length + 2)
                    column_letter = get_column_letter(column_index)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                    sheet.column_dimensions[column_letter].auto_size = True

                except:
                    pass
            # Устанавливаем фиксированную высоту и шрину для всех строк
            sheet.row_dimensions[row_index].auto_size = True
            sheet.column_dimensions[column_letter].auto_size = True

            # Вставляем предложенный способ подбора ширины столбцов
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


    # Функция для форматирования чисел и ячеек в таблице stock_plan
    def format_stock_plan(sheet):
        def format_stock_plan(sheet):
            # Создаем стиль для границ ячеек
            border_style = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        # Округляем числа и устанавливаем формат
                        cell.value = round(cell.value)
                        cell.number_format = '0'
                        if cell.value < -100:
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="000000")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            cell.font = Font(color="FF0000")
                        # Устанавливаем границы для ячейки
                        cell.border = border_style
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value < -100:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="000000")
                    elif cell.value < 0:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell.font = Font(color="FF0000")
                    cell.number_format = '0'


    output_file = f'Сводка планирования_{dates.iloc[0]}-{dates.iloc[-1]} {pc_name}.xlsx'
    counter = 1

    while os.path.exists(output_file):
        # Если файл с таким именем уже существует, добавляем порядковый номер
        output_file = f'Сводка планирования_{dates.iloc[0]}-{dates.iloc[-1]} {pc_name}_v{counter}.xlsx'
        counter += 1

    # Создание новых Excel-файлов
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        prod_plan.to_excel(writer, sheet_name='План Производства', index=False)
        autofit_columns(prod_plan, writer.sheets['План Производства'])
        format_stock_plan(writer.sheets['План Производства'])

        stock_plan.to_excel(writer, sheet_name='Остатки и Недопоставки', index=False)
        autofit_columns(stock_plan, writer.sheets['Остатки и Недопоставки'])
        format_stock_plan(writer.sheets['Остатки и Недопоставки'])

        resource_plan.to_excel(writer, sheet_name='Рабочие Центры', index=False)
        autofit_columns(resource_plan, writer.sheets['Рабочие Центры'])
        format_stock_plan(writer.sheets['Рабочие Центры'])

    if args.master and args.telegram:
        # Создание сводной таблицы с суммами отрицательных и положительных значений
        summary_table = pd.DataFrame()

        # Вычисление сумм отрицательных и положительных значений для каждой колонки
        for column in stock_plan.columns[2:]:  # Начинаем с третьей колонки, чтобы не включить дату и первую колонку
            positive_sum = stock_plan.loc[stock_plan[column] > 0, column].sum()
            negative_sum = stock_plan.loc[stock_plan[column] < 0, column].sum()
            production_sum = production_plan.loc[stock_plan[column] != -1, column].sum()
            summary_table.at['Остаток', column] = positive_sum
            summary_table.at['Недопоставка', column] = negative_sum


        # Создаем график остатков и сохраняем его в виде JPG
        fig, ax = plt.subplots()
        ax.set_xlabel('Дата')
        ax.set_ylabel('Сумма остатков')
        ax.set_title('График остатков по датам')
        plt.xticks(rotation=45)
        summary_table.loc['Остаток'].plot(kind='line', ax=ax, label='Остаток')
        plt.tight_layout()

        # Автоматически подбираем размер фигуры
        plt.gcf().autofmt_xdate()

        plt.savefig(f'остатки_{counter}.jpg', format='jpg')
        plt.close(fig)

        # Создаем график недопоставок и сохраняем его в виде JPG
        fig, ax = plt.subplots()
        ax.set_xlabel('Дата')
        ax.set_ylabel('Сумма недопоставок')
        ax.set_title('График недопоставок по датам')
        plt.xticks(rotation=45)
        summary_table.loc['Недопоставка'].plot(kind='bar', ax=ax, label='Недопоставка')
        plt.tight_layout()

        # Автоматически подбираем размер фигуры
        plt.gcf().autofmt_xdate()

        plt.savefig(f'недопоставки_{counter}.jpg', format='jpg')
        plt.close(fig)

        # Опционально: закрыть графическое окно, если вы используете интерактивное окружение
        plt.close(fig)
        print("aa")
        files=[f'остатки_{counter}.jpg', f'недопоставки_{counter}.jpg']
        send_media_group(files)
        send_file(output_file)







